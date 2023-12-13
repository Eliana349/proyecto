from django.shortcuts import render
from django.shortcuts import redirect
from django.contrib.auth import login
from django.contrib.auth import authenticate
from django.contrib import messages
from django.contrib.auth import logout
from .forms import *
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from core.models import *
from django.core.mail import send_mail
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.decorators import user_passes_test
from django.urls import reverse
from openpyxl import Workbook
from core.models import Product
from core.models import loyalty
from openpyxl.styles import NamedStyle
from django.shortcuts import get_object_or_404
from decimal import Decimal
from django.db import transaction






def generate_excel_report_carrito(request):
    # Crear un nuevo libro de Excel y una hoja de cálculo
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Escribir encabezados en la primera fila
    headers = ['Nombre de usuario', 'Fecha de inicio', 'Fecha de finalización', 'Elementos seleccionados', 'Precio Total']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Obtener datos de la base de datos y escribir en el archivo Excel
    alquileres = Carrito.objects.all()
    for row_num, alquiler in enumerate(alquileres, 1):
        data = [
            str(alquiler.nombre_usuario), alquiler.date_start, alquiler.date_finish,
            alquiler.elementos_alquilar, alquiler.precio_total
        ]

        # Escribir los datos en las celdas de Excel
        for col_num, value in enumerate(data, 1):
            sheet.cell(row=row_num + 1, column=col_num, value=value)

    # Crear la respuesta HTTP con el archivo adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_carrito.xlsx'
    workbook.save(response)

    return response




def generate_excel_report_loyalty(request):
    # Crear un nuevo libro de Excel y una hoja de cálculo
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Escribir encabezados en la primera fila
    headers = ['Nombres y apellidos', 'Correo electrónico', 'Número de teléfono', 'Tipo de PQRSD',
               'Fecha de incidente', 'Descripción detallada', 'Producto o servicio', 'Número de radicado',
               'Cómo prefiere ser contactad@']

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Obtener datos de la base de datos y escribir en el archivo Excel
    fidelizaciones = loyalty.objects.all()
    for row_num, fidelizacion in enumerate(fidelizaciones, 1):
        data = [
            fidelizacion.full_name, fidelizacion.email, fidelizacion.phone, fidelizacion.type_pqrsd,
            fidelizacion.incident_date, fidelizacion.detailed_description, fidelizacion.product_or_services_name,
            fidelizacion.filing_number, fidelizacion.preference_contact,
        ]

        # Escribir los datos en las celdas de Excel
        for col_num, value in enumerate(data, 1):
            sheet.cell(row=row_num + 1, column=col_num, value=value)

    # Crear la respuesta HTTP con el archivo adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_fidelizaciones.xlsx'
    workbook.save(response)

    return response




def generate_excel_report_cotizacion(request):
    # Crear un nuevo libro de Excel y una hoja de cálculo
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Escribir encabezados en la primera fila
    headers = ['Nombre Completo', 'Correo Electrónico', 'Número de Teléfono', 'Tipo de Evento', 'Fecha del Evento',
               'Duración del Evento (horas)', 'Sede del Evento', 'Número del salón', 'Cantidad de Invitados', 'Menú',
               'Cantidad de Menús Infantiles', 'Entradas Adicionales', 'Comentarios Adicionales', 'Servicios Adicionales',
               'Servicios Requeridos del Paquete Base', 'Tematica', 'Necesidad especial']

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Obtener datos de la base de datos y escribir en el archivo Excel
    cotizaciones = Cotizacion.objects.all()
    for row_num, cotizacion in enumerate(cotizaciones, 1):
        data = [
            cotizacion.name, cotizacion.email, cotizacion.phone_number, cotizacion.event_type,
            cotizacion.event_date, cotizacion.event_duration, cotizacion.event_location,
            cotizacion.salon_number, cotizacion.number_of_guests, cotizacion.menu,
            cotizacion.childrens_menu, cotizacion.additional_entries, cotizacion.additional_comments,
            cotizacion.additional_services, cotizacion.required_services, cotizacion.theme, cotizacion.special_need
        ]

        # Escribir los datos en las celdas de Excel
        for col_num, value in enumerate(data, 1):
            sheet.cell(row=row_num + 1, column=col_num, value=value)

    # Crear la respuesta HTTP con el archivo adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_cotizaciones.xlsx'
    workbook.save(response)

    return response


def generate_excel_report_pse(request):
    # Crear un nuevo libro de Excel y una hoja de cálculo
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Escribir encabezados en la primera fila
    headers = ['Tipo de persona', 'Seleccione su Banco', 'Nombre Completo', 'Tipo de Identificación', 'Numero de identificacion', 'Correo electronico', 'Numero de telefono']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Obtener datos de la base de datos y escribir en el archivo Excel
    pse_accounts = PSE.objects.all()
    for row_num, pse_account in enumerate(pse_accounts, 1):
        data = [
            str(pse_account.type_person), str(pse_account.select_bank), pse_account.full_name,
            str(pse_account.type_id), str(pse_account.identification_number), pse_account.email,
            str(pse_account.phone_number)
        ]

        # Escribir los datos en las celdas de Excel
        for col_num, value in enumerate(data, 1):
            sheet.cell(row=row_num + 1, column=col_num, value=value)

    # Crear la respuesta HTTP con el archivo adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_pse.xlsx'
    workbook.save(response)

    return response



def es_admin(user):
    return user.is_authenticated and user.is_staff

def redireccionar_admin(request):
    admin_url = reverse('admin:index')
    return redirect(admin_url)

def custom_excel_report(request):
    # Crear un libro de trabajo y obtener la hoja activa
    workbook = Workbook()
    sheet = workbook.active

    date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD HH:MM:SS')

    # Añadir encabezados a la hoja de cálculo
    headers = ['ID', 'Nombre', 'Producto', 'Precio', 'Cantidad', 'Fecha de Creación']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Obtener datos del modelo TipoDeProducto
    tipo_productos = TipoDeProducto.objects.all()

    # Llenar la hoja de cálculo con los datos del modelo TipoDeProducto
    for row_num, tipo_producto in enumerate(tipo_productos, 2):
        sheet.cell(row=row_num, column=1, value=tipo_producto.id)
        sheet.cell(row=row_num, column=2, value=tipo_producto.nombre)
        sheet.cell(row=row_num, column=3, value=str(tipo_producto.product))
        sheet.cell(row=row_num, column=4, value=tipo_producto.precio)
        sheet.cell(row=row_num, column=5, value=tipo_producto.cantidad)
        
        created_at_formatted = tipo_producto.created_at.strftime('%Y-%m-%d %H:%M:%S')
        sheet.cell(row=row_num, column=6, value=created_at_formatted).style = date_style

    # Crear una respuesta de Django con el contenido del libro de trabajo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=custom_excel_report.xlsx'
    workbook.save(response)

    return response


def generate_excel_report(request):
    # Crear un nuevo libro de Excel y una hoja de cálculo
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Escribir encabezados en la primera fila
    headers = ['Nombres', 'Apellidos', 'Correo electrónico', 'Número de celular', 'Género','Fecha de evento', 'Hora inicial', 'Hora final', 'Tematica',  'Necesidad especial', 'Tipo de evento', 'Sede', 'Salón']
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Obtener datos de la base de datos y escribir en el archivo Excel
    reservations = Reserva.objects.all()
    for row_num, reservation in enumerate(reservations, 1):
        data = [
            reservation.name, reservation.lastname, reservation.email,
            str(reservation.phone),  # Convertir PhoneNumber a cadena
            reservation.gender, reservation.event_date, reservation.event_start_time,
            reservation.end_time_of_the_event, reservation.theme, reservation.description,
            reservation.special_need, reservation.eventType, reservation.campus, reservation.lounge
        ]

        # Escribir los datos en las celdas de Excel
        for col_num, value in enumerate(data, 1):
            sheet.cell(row=row_num + 1, column=col_num, value=value)

    # Crear la respuesta HTTP con el archivo adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_reservas.xlsx'
    workbook.save(response)

    return response






def index(request):
    return render(request, 'home.html',{
    })


def login_view(request):    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username=username, password=password)
        
        if user:
            login(request, user)
            messages.success(request, 'Bienvenid@ {}'.format(user.username))
            return redirect('home2')
        else: 
            messages.error(request, 'Usuario o contraseña incorrectos')
    return render(request, 'login.html',{
    })



def register(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            names = form.cleaned_data.get('Names')
            surnames = form.cleaned_data.get('surnames')
            username = form.cleaned_data.get('username')
            email = form.cleaned_data.get('email')
            phone = form.cleaned_data.get('phone')
            password = form.cleaned_data.get('password')
            
            
            user = User.objects.create_user(username=username, email=email, password=password)
            user.first_name = names
            user.last_name = surnames
            user.save()
            
            login(request, user)
            return redirect('home2')
    else:
        form = RegisterForm()

    return render(request, 'register.html', {
        'form': form
    })

@login_required 
def home_two(request):
    return render(request, 'home2.html')


@login_required
def obtener_fechas_reservas_anteriores():
    reservas_anteriores = Reserva.objects.filter(event_date__lt=timezone.now().date())
    return [reserva.event_date for reserva in reservas_anteriores]

def reservas_view(request):
    reserva = get_object_or_404(Reserva, cotizacion_id=cotizacion_id)

    if request.method == 'POST':
        contact_form = formularioReserva(data=request.POST)

        if contact_form.is_valid():
            reserva = contact_form.save(commit=False)
            reserva.event_date = contact_form.cleaned_data['event_date']
            reserva.event_start_time = contact_form.cleaned_data['event_start_time']
            reserva.end_time_of_the_event = contact_form.cleaned_data['end_time_of_the_event']
            reserva.save()
            return redirect('tarjeta')

        else:
            messages.error(request, 'Hubo un error en el formulario')
            messages.success(request, 'Datos no permitidos')


    else:
        contact_form = formularioReserva()

    return render(request, 'reservas.html', {'form': contact_form})

@login_required
def pse_view(request, cotizacion_id):
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)

    if request.method == 'POST':
        pse_form = formularioPSE(request.POST)
        if pse_form.is_valid():
            pse_instance = pse_form.save(commit=False)
            pse_instance.cotizacion = cotizacion
            pse_instance.save()

            # Envía un correo electrónico de confirmación
            subject = 'Confirmación de Pago PSE'
            message = '¡Gracias por tu pago! Tu transacción ha sido procesada con éxito.'
            from_email = 'casaberrio23@gmail.com'  # Cambia esto al correo desde el cual deseas enviar
            recipient_list = [pse_instance.email]  # Ajusta según el nombre real de tu campo de correo electrónico

            send_mail(subject, message, from_email, recipient_list, fail_silently=False)

            # Redirige a la página de inicio o a donde desees después de un envío exitoso
            return redirect('home2')
        else:
            for field, error in pse_form.errors.items():
                messages.error(request, f"{field}: {error}")

    else:
        # Inicializa el formulario PSE con los datos de la cotización
        pse_form = formularioPSE(initial={
            'cotizacion_id': cotizacion.id,
            'full_name': cotizacion.name,
            'email': cotizacion.email,
            'phone_number': cotizacion.phone_number,
            'Total_value': cotizacion.valor_total,
            # Agrega otros campos según sea necesario
        })

    return render(request, 'PSE.html', {'form': pse_form, 'cotizacion_id': cotizacion_id})

@login_required
def tajetacd_view(request):
    contact_form = formularioTarjetaDeCD()
    
    if request.method == 'POST':
        contact_form = formularioTarjetaDeCD(data=request.POST)
        
        if contact_form.is_valid():
            contact_form.save()
            return redirect('home2')
            
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
        
    return render(request,  'tarjetacd.html', {'form':contact_form})


def fidelizacion_view(request):
    contact_form = formularioFedelizacion()
    
    if request.method == 'POST':
        contact_form = formularioFedelizacion(data=request.POST)
        
        if contact_form.is_valid():
            contact_form.save()
            return redirect('home')
            
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
        
    return render(request,  'pqrs.html', {'form':contact_form})


@login_required
def inventario_view(request):
    contact_form = formularioInventario()
    
    if request.method == 'POST':
        contact_form = formularioInventario(data=request.POST)
        
        if contact_form.is_valid():
            contact_form.save()
            return redirect('home2')
            
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
        
    return render(request,  'pqrs.html', {'form':contact_form})


@login_required
def inventario_view(request):
    contact_form = formularioInventario()
    
    if request.method == 'POST':
        contact_form = formularioInventario(data=request.POST)
        
        if contact_form.is_valid():
            contact_form.save()
            return redirect('home2')
            
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
        
    return render(request,  'pqrs.html', {'form':contact_form})
   


@login_required
def alquiler_view(request):
    contact_form_tipoP = TipoDeProducto.objects.all()
    contact_formcategoria = Category.objects.all()
    contact_formtipo = Product.objects.all()
    contact_form = formularioTipo(request.POST or None)
    contact_form_carrito = formularioCarrito(request.POST or None)

    if contact_form_carrito.is_valid():
        with transaction.atomic():
            # Restar la cantidad de productos alquilados del inventario
            for opcion in request.session.get('opciones_alquiler', []):
                _, tipo, cantidad, _ = opcion.split('  :  ')
                tipo_producto = TipoDeProducto.objects.get(nombre=tipo)
                tipo_producto.cantidad -= int(cantidad)
                tipo_producto.save()

            # Guardar el formulario del carrito
            contact_form_carrito.save()

            # Limpiar la sesión después de un alquiler exitoso
            request.session.pop('opciones_alquiler', None)
            request.session.pop('total_alquiler', None)

        return redirect('psee')

    t = Decimal(request.session.get('total_alquiler', 0))

    if contact_form.is_valid(): 
        Tipo = request.POST.get('option1')
        number = contact_form.cleaned_data.get('cantidad')
        filterr = TipoDeProducto.objects.get(nombre=Tipo)
        price = filterr.precio
        precio = number * price
        producto = filterr.product
        amount = filterr.cantidad

        if number > amount:
            return HttpResponse('<h1 style="color:red; font-size: 50px">CANTIDAD NO DISPONIBLE</h1>')

        opciones_nuevas = f'{producto}  :  {Tipo}  :  {number}  :  ${precio}'
        t += precio
        total = f'$ {t}'
        opciones_guardadas = request.session.get('opciones_alquiler', [])
        opciones_guardadas.append(opciones_nuevas)
        request.session['opciones_alquiler'] = opciones_guardadas
        request.session['total_alquiler'] = float(t)
        opciones_en_carrito = '\n'.join(opciones_guardadas)
        contact_form_carrito = formularioCarrito({'elementos_alquilar': opciones_en_carrito, 'precio_total': total})

    return render(request, 'alquiler.html', {
        'cantidad': contact_form, 
        'Carrito': contact_form_carrito,
        'products': contact_formtipo,
        'Categoria': contact_formcategoria,
        'Tipo': contact_form_tipoP,
    })
    
def limpiar_sesion(request):
    if 'opciones_alquiler' in request.session:
        del request.session['opciones_alquiler']
    if 'total_alquiler' in request.session:
        del request.session['total_alquiler']
    return HttpResponse("Sesión limpiada")    
        

@login_required
def alquiler_productsView(request):
    product = Product.objects.all()
    category = Category.objects.all()
    tipo = TipoDeProducto.objects.all()
    return render(request, 'alquiler_products.html', {
        'producto':product,
        'categoria': category,
        'tipo':tipo
    })
    
    
@login_required
def alquiler_historialView(request):
    car = Carrito.objects.all()
    return render(request, 'alquiler_historial.html', {
        'carro': car,
    })
    
    

@login_required
def ppsse_view(request):
    contact_form = formularioPSE()

    if request.method == 'POST':
        contact_form = formularioPSE(data=request.POST)
        
        if contact_form.is_valid():
            # Guardar el formulario
            contacto = contact_form.save()

            # Envía un correo electrónico de confirmación
            subject = 'Confirmación de Pago PSE'
            message = '¡Gracias por tu pago! Tu transacción ha sido procesada con éxito.'
            from_email = 'casaberrio23@gmail.com'  # Cambia esto al correo desde el cual deseas enviar
            recipient_list = [contacto.email]  # Ajusta según el nombre real de tu campo de correo electrónico

            send_mail(subject, message, from_email, recipient_list, fail_silently=False)

            # Redirige a la página de inicio o a donde desees después de un envío exitoso
            return redirect('home2')
            
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
        
    return render(request, 'ppsse.html', {'form': contact_form})

@login_required
def productos(request):
    return render(request, 'productos.html',{
    })

def nosotros(request):
    return render(request, 'nosotros.html',{
    })

def crear_cotizacion(request):
    form = CotizacionForm(request.POST or None)
    cotizaciones = Cotizacion.objects.all()
    if request.method == 'POST':
        form = CotizacionForm(request.POST)
        if form.is_valid():
            form.instance.state = 'En espera'
            valor_total = 0
            duracion = form.cleaned_data['event_duration']
            if duracion == 4: 
                valor_total += 3500000
            elif duracion == 5:
                valor_total += 4000000
            elif duracion == 6:
                valor_total += 4500000
            elif duracion == 7:
                valor_total += 5000000

            salon = form.cleaned_data['salon_number']
            valor_total += 1000000

            cantidad_invitados = form.cleaned_data['number_of_guests']
            if 40 <= cantidad_invitados <= 69:
                valor_total += 1000000
                valor_total += 800000 #valor adicional del menu
                valor_total += 200000 #valor adicional de las entradas
            elif  70 <= cantidad_invitados <= 99:
                valor_total += 1500000
                valor_total += 1000000 #valor adicional del menu
                valor_total += 400000 #valor adicional de las entradas
            elif  100 <= cantidad_invitados <= 129:
                valor_total += 2000000
                valor_total += 1200000 #valor adicional del menu
                valor_total += 600000 #valor adicional de las entradas
            elif  130 <= cantidad_invitados <= 159:
                valor_total += 2500000
                valor_total += 1500000 #valor adicional del menu
                valor_total += 800000 #valor adicional de las entradas
            elif  160 <= cantidad_invitados <= 180:
                valor_total += 3000000
                valor_total += 1800000 #valor adicional del menu
                valor_total += 1000000  #valor adicional de las entradas

            paquete_base = form.cleaned_data['required_services']
            for servicio in paquete_base:
                if servicio == 'Pastel':
                 valor_total += 300000
                elif servicio == 'Sonido Dj - Animación':
                    valor_total += 600000
                elif servicio == 'Sillas':
                    valor_total += 100000
                elif servicio == 'Centros de Mesa':
                    valor_total += 150000
                elif servicio == 'Fotografía Digital':
                    valor_total += 250000
                elif servicio == 'Mezcladores, hielos y gaseosas':
                    valor_total += 120000
                elif servicio == 'Coctel':
                    valor_total += 180000
                elif servicio == 'Servicio de meseros por 7 horas':
                    valor_total += 500000
                elif servicio == 'Champañas':
                    valor_total += 150000

            servicios_adicionales = form.cleaned_data['additional_services']
            for servicio in servicios_adicionales:
                if servicio == 'Decoración Escalera en Flores Naturales':
                 valor_total += 300000
                elif servicio == 'Kit Hora de Carnaval con Accesorios':
                    valor_total += 200000
                elif servicio == 'Fiesta Temática':
                    valor_total += 400000
                elif servicio == 'Otros - Filmación':
                    valor_total += 350000
                elif servicio == 'Foto Registro Enmarcado 40x50 cms':
                    valor_total += 150000
                elif servicio == 'Libro de Firmas':
                    valor_total += 125000
                elif servicio == 'Videos Retrospectivos':
                    valor_total += 400000
                elif servicio == 'Video Beam - Telón':
                    valor_total += 225000
                elif servicio == 'Whisky':
                    valor_total += 600000    
            else:
                valor_total += 0   
            cotizacion = form.save(commit=False)
            cotizacion.valor_total = valor_total
            cotizacion.save()   
            return redirect('reserva', cotizacion_id=cotizacion.id)
          
    return render(request, 'cotizaciones.html', {'form': form,'cotizacions': cotizaciones})

def cotizacion_vista(request):
    return render(request, 'cotizacion.html', {})

def reserva(request, cotizacion_id):
    cotizacion = get_object_or_404(Cotizacion, id=cotizacion_id)

    if request.method == 'POST':
        reserva_form = formularioReserva(request.POST)
        if reserva_form.is_valid():
            reserva = reserva_form.save(commit=False)
            reserva.cotizacion = cotizacion
            reserva.save()
            return redirect('PSE.html')

    else:
        reserva_form = formularioReserva(initial={
            'name': cotizacion.name,
            'email': cotizacion.email,
            'phone': cotizacion.phone_number,
            'eventType': cotizacion.event_type,
            'campus': cotizacion.event_location,
            'theme': cotizacion.theme,
            'special_need': cotizacion.special_need,
            'lounge': cotizacion.salon_number,
            'Total_value': cotizacion.valor_total,

        })

    return render(request, 'reservas.html', {'reserva_form': reserva_form, 'cotizacion': cotizacion})

    